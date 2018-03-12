import requests
from bs4 import BeautifulSoup
import openpyxl
import re
import time

'''
1st function call. Searches siccode.com for business name (from Excel)
and returns all <a> tags if there are multiple results. One of these urls
on the results_page will be the business in Madison Heights/MI.
'''
# biz_name is from spreadsheet
def get_site(biz_name):
	
	# returns all search results on 1 page
	search_url = r'https://siccode.com/en/business-list/' + biz_name
	
	# get page object, save as string
	results_page = requests.get(search_url)
	results_page_doc = results_page.text
	
	# create soup object
	soup = BeautifulSoup(results_page_doc, 'html.parser')
	
	# get all <a>s to use regex on
	results_links = soup.find_all('a')
	
	# return all soup <a> objects
	return results_links
	
'''
2nd function call. Loops thru all <a> links on search results page and 
uses regex to find link of business in Michigan (starts with 480..).
'''
# takes as argument all <a> links from business search results
def get_url(results_links):
	
	for link in results_links:
		
		# returns a string using a BS method
		url = link.get('href')
		
		# make regex object to find first(?) url that contains MI zip code 
		url_regex = r'.*48\d\d\d'
		
		try:
			
			# create match object
			mo = re.search(url_regex, url)
			
			# url string that points to business-specific page
			biz_url = mo.group()
			
			# might return None(?)
			return biz_url
		
		# if mo == None:
		except:
			pass
			
'''
3rd function call. Calls page specific to business and 
gets NAICS code on that page. May eventually merge this 
with 2nd function call.
'''
def call_url(biz_url):
	
	# calls business-specific page
	biz_page_url = r'http://siccode.com' + biz_url
	
	# converts entire page object to text
	biz_page = requests.get(biz_page_url)
	biz_page_doc = biz_page.text
	
	# gets all <a> text from page
	soup_2 = BeautifulSoup(biz_page_doc, 'html.parser')
	naics_a = soup_2.find_all('a')
	
	# for each element in <a> object
	for elem in naics_a:
		
		# change each one into text; move outside for loop?
		elem_doc = elem.text
		
		# gets 6-digit naics code and description after dash
		elem_regex = r'^\s(\d{6})\s-\s(.*)'
		
		# may return None; handle this downstream
		try:
		
			mo = re.search(elem_regex, elem_doc)
			code = mo.group(1)
			descr = mo.group(2)
			
			return code, descr
		
		except:
			pass

# main loop
def master():
	
	# create workbook & sheet(s) objects
	wb = openpyxl.load_workbook('biz.xlsx')
	sheet = wb.active
	
	max_col = sheet.max_column
	
	# create headers
	sheet.cell(row = 1, column = max_col + 1, value = 'naics code')
	sheet.cell(row = 1, column = max_col + 2, value = 'description')
	
	#loop through rows to get biz name
	for row_iter in range(2, sheet.max_row + 1):
		
		# get business names from Excel
		biz_name = sheet.cell(row = row_iter, column = 5).value
		
		# print record being looked up and print result on same line
		print('Looking up code for %s...' % (biz_name), end = '');
		
		# returns soup object of all results matching business name
		results_links = get_site(biz_name)
		
		# returns url segment specific to MI business 
		biz_url = get_url(results_links)
		
		try:
		
			# uses regex to get naics code from specific website
			# handle None type below
			code, descr = call_url(biz_url)
			
			# let user know on CLI (mostly for fun)
			print(' %s - %s' % (str(code), descr))
			
			# write results to sheet
			sheet.cell(row = row_iter, column = max_col + 1, value = code)
			sheet.cell(row = row_iter, column = max_col + 2, value = descr)
		
			# rest for a moment to avoid battering their servers
			time.sleep(5)
		
		# if biz_url is None
		except:
			
			sheet.cell(row = row_iter, column = max_col + 1, value = 'NOT FOUND')
			
			print(' %s' % 'NOT FOUND')
		
		
		# save temp file every 250 records
		if (row_iter % 251 == 0):
			
			wb.save('biz_temp_%i.xlsx' % (row_iter-1))
	
	wb.save('biz_coded.xlsx')
			
master()
		

